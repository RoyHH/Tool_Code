from openpyxl import load_workbook
import numpy as np

def excel_read(path):
    excel_open = load_workbook(path, data_only=True)  # 打开excel
    Sheet = excel_open.active  # 表示当前活跃的表，本案例中 当前活跃表为sheet1
    # 也可以使用 Sheet = excel_open.get_sheet_by_name('Sheet1') 来获取工作表1
    Row_Num = Sheet.max_row  # 读取excel行数
    Col_Num = Sheet.max_column  # 读取excel列数

    Dict_X = locals()  # locals():以字典类型返回当前位置的全部局部变量
    for i in range(Row_Num):
        ii = i+1
        dict_x = []
        for j in range(Col_Num):
            jj = j+1
            a = []
            a = str(Sheet.cell(row=ii, column=jj).value)
            # print(a)
            if a == "None":
                dict_x.append('\0')
            else:
                dict_x.append(a)
            # print(dict_x)
        Dict_X['dict_x' + str(i)] = np.array(dict_x)
        # print(Dict_X)
    for i in range(Row_Num):
        print(Dict_X['dict_x' + str(i)])

    print(Row_Num, Col_Num)
    print(Dict_X['dict_x1'][2])
    return Dict_X, Row_Num, Col_Num

if __name__ == '__main__':

    path = r'数据采集记录.xlsx'

    keypoint = []
    keypoint, Row_Num, Col_Num = excel_read(path)
    print('啥玩意=', keypoint['dict_x0'])
